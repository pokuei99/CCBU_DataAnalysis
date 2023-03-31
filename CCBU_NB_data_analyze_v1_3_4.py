from genericpath import exists
import os
import matplotlib
import numpy as np
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
import pandas as pd
from pandas import DataFrame
from pandas import ExcelWriter
from pandas import read_csv
from pandas import read_excel
from pandas import concat
import tkinter as tk
from tkinter import *
from matplotlib import pyplot as plt
import csv
import xlrd
import xlwt
from tkinter import messagebox
from tkinter import messagebox, filedialog
from openpyxl.utils.dataframe import dataframe_to_rows


dir=""
window = tk.Tk()
window.title("CCBU_THMNB_Analyzer_v1.3.4")
canvas = tk.Canvas(window, width=853, height=520, bg='black')#width=750, height=500
canvas.pack(expand=YES, fill=BOTH)
#gif1 = PhotoImage(file='.\\charizard_test.gif')
#canvas.create_image(0, 0, image=gif1, anchor=NW)

#choose folder  
def button_event1():
    global dir 
    if dir=="":
        dir = filedialog.askdirectory(initialdir='./')
    else:
        dir = filedialog.askdirectory()
    data_file.set(dir)

############ to be continued
def run_all():     
    #check data file
    if dir=="":
        messagebox.showinfo("Notification","Please select data file.")
        return
    else:
        entry = os.path.basename(dir)
        #print(entry)
        dirdir = os.listdir(dir)      # 列出資料夾檔案
        resultbook = dir + "\\" + entry + '_result.xlsx'
        sen_plot_path = dir + "\\" + entry + '_SEN-temp_plot.png'
        tcpu_plot_path = dir + "\\" + entry + '_TCPU-PKG-temp_plot.png'
        ssd_plot_path = dir + "\\" + entry + '_SSD-temp_plot.png'
        yoko_plot_path = dir + "\\" + entry + '_YOKOplot.png'
        pkgp_plot_path = dir + '\\' + entry + '_Power-Package Power_plot.png'
        tmm_rpm_path = dir + '\\' + entry + '_FAN_RPM_plot.png'
        filetype_tat = r"S.csv"
        filetype_yoko = r".xls"
        filetype_hwi = r".CSV"
        filetype_gmon = r".log"
        filetype_tmm = r".log"         

    #check acquiring data time
    if myentry.get()=="":
        messagebox.showinfo("Notification","Please type the acquiring data length.")
        return
    elif int(myentry.get())> 60:
        messagebox.showinfo("Notification","Acquiring data is too long.")
        return
    else:
        data_time_1 = myentry.get()
        data_time_2 = myentry_2.get()         
        data_length_1 = int(data_time_1)*60
        data_length_2 = int(data_time_2)*60 
    #####################################################
    #remove exsisted excel
    if os.path.exists(resultbook):
        os.remove(resultbook)
    if os.path.exists(pkgp_plot_path):
        os.remove(pkgp_plot_path)
    if os.path.exists(sen_plot_path):
        os.remove(sen_plot_path)
    if os.path.exists(tcpu_plot_path):
        os.remove(tcpu_plot_path)
    if os.path.exists(ssd_plot_path):
        os.remove(ssd_plot_path)
    if os.path.exists(yoko_plot_path):
        os.remove(yoko_plot_path)
    if os.path.exists(tmm_rpm_path):
        os.remove(tmm_rpm_path)
    
    #open new excel
    with pd.ExcelWriter(dir + '\\' + entry + '_result.xlsx',engine='openpyxl') as writer_start:   
        empty=pd.DataFrame([])
        empty.to_excel(writer_start, sheet_name='avg')
        #worksheet = writer.sheets['avg']
    avg_col_num = 0
    if CheckVar1.get() == 1: #TAT
        if any(File.startswith('PTAT') or File.startswith('TAT') for File in os.listdir(dir)):#可以補上startswith('TAT') 2022/07/15
            [_] = [fname for fname in dirdir if fname.startswith('PTAT') or fname.startswith('TAT')] #可以補上startswith('TAT') 2022/07/15  
            try:
                data_tat = pd.DataFrame(pd.read_csv(os.path.join(dir, _)))    
                data_tat.set_index('Time', inplace=True)
            except:
                messagebox.showinfo("Notification","TAT file log error, please check the raw data")
                return
            
            col_small = [col for col in data_tat.columns if data_tat[col][1]=='Small Core']
            col_small_freqMHz = [w.replace('Core Type','Frequency(MHz)') for w in col_small]
            col_small_freq = [w.replace('Core Type','Frequency') for w in col_small]
            try:
                data_tat.drop(columns=col_small_freq, inplace=True)
            except:
                pass
            try:
                data_tat.drop(columns=col_small_freqMHz, inplace=True)
            except:
                pass

            tat_sen=[]
            tat_ssd=[]
            tatresult_freq= data_tat.filter(regex="CPU\d+-Frequency",axis=1)
            tatresult_freq['freq_avg'] = tatresult_freq.mean(axis=1)
            tatresult_pch = data_tat.filter(regex="PCH-Max Dts Temperature",axis=1)
            tatresult_dts = data_tat.filter(regex="MSR Package Temperature",axis=1)
            tatresult_dts2 = data_tat.filter(regex="MMIO Package Temperature",axis=1)
            tatresult_power = data_tat.loc[:,data_tat.columns.str.contains("Power-EWMA|Power-IA|Power-Integrated Graphics Power|Power-GT|Power-Rest of Package Power")]
            tatresult_pkgp = data_tat.loc[:,data_tat.columns.str.contains("Power-Package Power")]
            tatresult_temperature = data_tat.loc[:,data_tat.columns.str.contains("TCPU-CPU-temp|TCPU-PKG-temp")]
            tatresult_plx = data_tat.loc[:,data_tat.columns.str.contains("TCC Offset|MSR Power Limit_1 Power|MSR Power Limit_2 Power|MMIO Power Limit_1 Power|MMIO Power Limit_2 Power")]
            tatresult_thmevent = data_tat.loc[:,data_tat.columns.str.contains("Clip Reason")]
            #filter sen and ssd in columns
            for col_name in data_tat.columns:
                if CheckVar1_1.get() == 1:
                    if "DG0" in col_name:
                        tat_ssd.append(col_name) #不額外增加變數直接把項目加到後面
                if "SEN" in col_name:
                    tat_sen.append(col_name)
                if "SSD" in col_name:
                    tat_ssd.append(col_name)
                
            tatresult_sen = data_tat.loc[:,tat_sen]
            tatresult_ssd = data_tat.loc[:,tat_ssd]
            
            tatresult_allpower = tatresult_power.join(tatresult_pkgp)
            try:
                tatresult_allpower.plot(title='PTAT Power(W)',  #圖表標題
                            xlabel='t',  #x軸說明文字
                            ylabel='W',  #y軸說明文字
                            legend=True,  # 是否顯示圖例 #圖例多個29可以把它debug掉
                            figsize=(20, 10))  # 圖表大小
                plt.savefig(dir + '\\' + entry + '_Power-Package Power_plot')
            except:
                pass
            try:
                tatresult_temperature.plot(title='TCPU-PKG-temp(Degree C)',  #圖表標題
                        xlabel='t',  #x軸說明文字
                        ylabel='Degree C',  #y軸說明文字
                        legend=True,  # 是否顯示圖例 #圖例多個29可以把它debug掉
                        figsize=(20, 10))  # 圖表大小 
                plt.savefig(dir + '\\' + entry + '_TCPU-PKG-temp_plot')
            except:
                pass
            try:
                tatresult_sen.plot(title='SEN-temp(Degree C)',  #圖表標題
                        xlabel='t',  #x軸說明文字
                        ylabel='Degree C',  #y軸說明文字
                        legend=True,  # 是否顯示圖例 #圖例多個29可以把它debug掉
                        figsize=(20, 10))  # 圖表大小 
                plt.savefig(dir + '\\' + entry + '_SEN-temp_plot')
            except:
                pass
            try:
                tatresult_ssd.plot(title='SSD-temp(Degree C)',  #圖表標題
                        xlabel='t',  #x軸說明文字
                        ylabel='Degree C',  #y軸說明文字
                        legend=True,  # 是否顯示圖例 #圖例多個29可以把它debug掉
                        figsize=(20, 10))  # 圖表大小 
                plt.savefig(dir + '\\' + entry + '_SSD-temp_plot')
            except:
                pass
            # tat_ssd = [col_name for col_name in data_tat.columns if data_tat.columns.str.contains("SSD")]
            #print(tat_sen)
            #7/19 work till this, there still have some bug and unsolved problem these comments are still need

        
            tatlog_freq_avg = (tatresult_freq[-data_length_1:]).mean().loc[['freq_avg']].round(1)
            tatlog_pch_avg = (tatresult_pch[-data_length_1:]).mean().round(1)
            tatlog_dts_avg = (tatresult_dts[-data_length_1:]).mean().round(1)
            #tatlog_dts_max = (tatresult_dts[-data_length:]).max().round(1)
            tatresult_allpower_avg = (tatresult_allpower[-data_length_1:]).max().round(2)
            tatresult_temperature_mode = (tatresult_temperature[:]).mode(axis=0,numeric_only=True).T
            tatresult_sen_mode = (tatresult_sen[:]).mode(axis=0,numeric_only=True).T
            tatresult_ssd_mode = (tatresult_ssd[:]).mode(axis=0,numeric_only=True).T
            tat_last_avg = pd.concat([tatlog_pch_avg,tatlog_dts_avg,tatresult_allpower_avg,tatlog_freq_avg,tatresult_temperature_mode,tatresult_sen_mode,tatresult_ssd_mode])
            tat_last_raw = pd.concat([tatresult_freq,tatresult_pch,tatresult_dts,tatresult_dts2,tatresult_power,tatresult_pkgp,tatresult_temperature,tatresult_sen,tatresult_ssd,tatresult_plx,tatresult_thmevent],axis=1)
            ################################################################
            
            ############################################################
            
            #book = load_workbook(resultbook)
            with pd.ExcelWriter(resultbook, mode="a", engine='openpyxl',if_sheet_exists="overlay") as writer:
                
                tat_last_avg.to_excel(writer, sheet_name='avg', startcol = avg_col_num, header=False)
                tat_last_raw.to_excel(writer, sheet_name="TAT_raw")
                #writer.workbook = book
                #tat_last_avg.to_excel(writer, sheet_name='avg', startrow= writer.sheets.max_row+1, header=False)
                #writer.sheets = {ws.title:ws for ws in book.worksheets}#20230217 modified still have some bug
                #tat_last_avg.to_excel(writer, sheet_name='avg',startrow=writer.sheets['avg'].max_row+1, header=False)
                #tat_last_raw.to_excel(writer, sheet_name='TAT_raw')
            avg_col_num = avg_col_num + 2
        else:
            messagebox.showinfo("Notification","TAT data file does not exsist.")
            return
            
    # if CheckVar1.get() == 2:
    #     messagebox.showinfo("Notification","Not ready, bro.")
    #     return
    
    if CheckVar2.get() == 1: #YOKOGAWA
        if any(File.endswith(filetype_yoko) for File in os.listdir(dir)):
            #check channel
            
            # if entry_ch.get()=="":
            #     messagebox.showinfo("Notification","Please tell me how many channels.")
            #     return
            # elif int(entry_ch.get())> 60:
            #     messagebox.showinfo("Notification","Check your channel again, bastard.")
            #     return
            #else:
                #num_ch = int(entry_ch.get())
            #num_ch = int(0)        
            [_] = [fname for fname in dirdir if fname.endswith(filetype_yoko)]    
            yokolog_path = os.path.join(dir, _)
            yokolog_path2 = pd.DataFrame(pd.read_excel(yokolog_path))
            #print(yokolog_path2.iat[29 , 4])
            num_ch = 0
            for item in range(61):
                try:
                    if "CH" in yokolog_path2.iat[29 , 2 + item]:
                        num_ch += 1
                except IndexError:
                    break
            #print(num_ch)
            if num_ch > 61:
                messagebox.showinfo("Notification","What the hell ? Check your channel again.")
                return
            datatime=pd.DataFrame(yokolog_path2.iloc[33:,1] )
            ###################開始繪圖
            datatime.columns=['Time']
            yokolog_path3 = yokolog_path2.iloc[33:, 3:(num_ch+3)]
            yokolog_path3.columns = yokolog_path2.iloc[29, 3:(num_ch+3)]
            yokolog_I_V = yokolog_path3.iloc[:, -2:]
            yokolog_path3['System_Power'] = yokolog_I_V.prod(axis=1)
            
            yokolog_path3.index = datatime['Time']
            try:
                yokolog_path3.plot(title='Channal-temp(Degree C)',  #圖表標題
                    xlabel='t',  #x軸說明文字
                    ylabel='Degree C',  #y軸說明文字
                    legend=True,  # 是否顯示圖例 #圖例多個29可以把它debug掉
                    figsize=(20, 10))  # 圖表大小 
                plt.savefig(dir + '\\' + entry + '_YOKOplot')
            except:
                pass
            yoko_last_avg = yokolog_path3.iloc[-data_length_2:].mean().round(1)#round 取小數點後一位 another iloc 好像要換成 loc
            yoko_last_avg = yoko_last_avg[yoko_last_avg>0]#取欄位中大於0的值來取平均
            yoko_last_raw = yokolog_path3

            with pd.ExcelWriter(resultbook, mode="a", engine='openpyxl',if_sheet_exists="overlay") as writer:
                yoko_last_avg.to_excel(writer, sheet_name='avg', startcol = avg_col_num, header=False)
                yoko_last_raw.to_excel(writer, sheet_name="YOKO_raw")
            avg_col_num = avg_col_num + 2
            """ book = load_workbook(resultbook)
            with pd.ExcelWriter(resultbook, engine='openpyxl') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                yoko_last_avg.to_excel(writer, sheet_name='avg',startcol=(writer.sheets['avg'].max_column+2), header=False)
                yoko_last_raw.to_excel(writer, sheet_name='YOKO_raw') """
            
        else:
            messagebox.showinfo("Notification","Yokogawa data file does not exsist.")
            return 
    if CheckVar3.get() == 1: #TMM
        test_except = True
        # if os.path.exists(resultbook):
        #     os.remove(resultbook)
        if any(File.endswith(filetype_tmm) and File.startswith('ThermalMonitor') for File in os.listdir(dir)):
            [_] = [fname for fname in dirdir if (fname.endswith(filetype_tmm) and fname.startswith('ThermalMonitor') )]
            tmm_path = os.path.join(dir, _)
            filetype_tmmcsv = r"tmmresult.csv"
            tmm_txt = open(tmm_path, "r",encoding="utf-8").readlines()[15:]#指定讀取第15行到最後一行???????????
            tmm_txt2 = open(tmm_path, "r", encoding="utf-8").read()
            stripped = (line.strip(' ') for line in tmm_txt)
            lines = (line.split() for line in stripped if line)
            ###############################################
            if os.path.exists(dir + '\\' + entry + '_tmmresult.csv'):#if temporary result exist then remove it
                os.remove(dir + '\\' + entry + '_tmmresult.csv')
            with open(dir + '\\' + entry + '_tmmresult.csv', 'w') as out_file:#將log檔中內容寫進excel
                writer = csv.writer(out_file)
                writer.writerows(lines)
                dirdir = os.listdir(dir)

                for u in dirdir:
                    #print(dirdir)
                    if u.endswith(filetype_tmmcsv):

                        tmmlog_path2 = os.path.join(dir, u)
                        try:
                            tmmlog_path3 = DataFrame(read_csv(tmmlog_path2, skiprows=1, header=None))#*
                        except:
                            messagebox.showinfo("Notification", "The version of TMM has changed, it can only produce a temporary file")
                            test_except = False
                            break

                        if "uP1905(2)" in tmm_txt2:
                            
                            tmmlog_path4 = tmmlog_path3.iloc[:, [
                                2,3, 5, 6, 9, 11, 13, 14, 16, 17, 18, 19, 20, 21, 22, 24]]#add battery temperature

                            tmmlog_path4.columns = ['Time','DTS', 'TS0R', 'TS0L', 'PPWR', 'FAN1DUTY', 'FAN1RPM', 'FAN2DUTY', 'FAN2RPM', 'uP1905(1)CH1', 'uP1905(1)CH2', 'uP1905(1)CH3', 'uP1905(2)CH1', 'uP1905(2)CH2', 'uP1905(2)CH3','Battery_Temperature']
                            
                            # tmmlog_column_temperature = ['DTS', 'TS0R', 'TS0L', 'uP1905(1)CH1', 'uP1905(1)CH2', 
                            #                             'uP1905(1)CH3', 'uP1905(2)CH1', 'uP1905(2)CH2', 'uP1905(2)CH3']
                            # tmmlog_column_power = ['PPWR']
                            # tmmlog_column_fan = ['FAN1DUTY','FAN1RPM', 'FAN2DUTY', 'FAN2RPM']
                            #tmmlog_column_rpm = ['FAN1RPM', 'FAN2RPM']

                        elif "APL6012(1)" in tmm_txt2:
                            
                            tmmlog_path4 = tmmlog_path3.iloc[:, [
                                2,3, 5, 6, 9, 11, 13, 14, 16, 26, 27, 28, 29, 30, 31, 33]] #從轉好的excel已經組成dataframe

                            tmmlog_path4.columns = ['Time','DTS', 'TS0R', 'TS0L', 'PPWR', 'FAN1DUTY', 'FAN1RPM', 'FAN2DUTY', 'FAN2RPM', 'APL6012(1)CH10', 'APL6012(1)CH11', 'APL6012(1)CH12', 'APL6012(1)CH13', 'APL6012(1)CH14', 'APL6012(1)CH15','Battery_Temperature']
                            
                            # tmmlog_column_temperature = ['DTS', 'TS0R', 'TS0L', 'uP1905(1)CH1', 'uP1905(1)CH2', 
                            # tmmlog_column_temperature = ['DTS', 'TS0R', 'TS0L', 'APL6012(1)CH10', 'APL6012(1)CH11', 'APL6012(1)CH12', 'APL6012(1)CH13', 'APL6012(1)CH14', 'APL6012(1)CH15']
                            # tmmlog_column_power = ['PPWR']
                            # tmmlog_column_fan = ['FAN1DUTY','FAN1RPM', 'FAN2DUTY', 'FAN2RPM']
                            #tmmlog_column_rpm = ['FAN1RPM', 'FAN2RPM']
                        else:
                            messagebox.showinfo("Warning", "TMM sensor format error")
                            return
                        # tmmlog_temperature = tmmlog_path4.loc[1:,tmmlog_column_temperature]
                        # tmmlog_power = tmmlog_path4.loc[1:,tmmlog_column_power]
                        # tmmlog_fan = tmmlog_path4.loc[1:,tmmlog_column_fan]
                        # tmmlog_index = tmmlog_path3.iloc[:,[2]]
                        # tmmlog_index.columns = ['Time']
                        tmmlog_path4.set_index('Time' , inplace = True)
                        tmm_last_avg = tmmlog_path4.iloc[-data_length_1:].mean().round(1)
                        tmm_last_raw = tmmlog_path4
                        tmm_rpm=[]
                        for col_name in tmm_last_raw.columns:
                            if "RPM" in col_name:
                                tmm_rpm.append(col_name)
                        tmm_rpm_result = tmm_last_raw.loc[:,tmm_rpm]
                        try:
                            tmm_rpm_result.plot(title='FAN RPM',  #圖表標題
                                            xlabel='t',  #x軸說明文字
                                            ylabel='RPM',  #y軸說明文字
                                            legend=True,  # 是否顯示圖例 #圖例多個29可以把它debug掉
                                            figsize=(20, 10))  # 圖表大小 
                            plt.savefig(dir + '\\' + entry + '_FAN_RPM_plot')
                        except:
                            pass
                        #接著輸出raw data以及取平均
                        with pd.ExcelWriter(resultbook, mode="a", engine='openpyxl',if_sheet_exists="overlay") as writer:
                            tmm_last_avg.to_excel(writer, sheet_name='avg', startcol = avg_col_num, header=False)
                            tmm_last_raw.to_excel(writer, sheet_name="TMM_raw")
                        avg_col_num = avg_col_num + 2
                        #book = load_workbook(resultbook)
                        """ with pd.ExcelWriter(resultbook, engine='openpyxl') as writer:
                            writer.book = book
                            writer.sheets = {ws.title: ws for ws in book.worksheets}
                            tmm_last_avg.to_excel(writer, sheet_name='avg',startcol=(writer.sheets['avg'].max_column+2), header=False)
                            tmm_last_raw.to_excel(writer, sheet_name='TMM_raw') """
                    
                    # else :
                    #     messagebox.showinfo("Warning", "My fault and let me debug")
            if  test_except == True:
                os.remove(dir + '\\' + entry + '_tmmresult.csv')
            else:
                pass
        else:
            messagebox.showinfo("Notification","TMM data file does not exsist.")
            return

    if CheckVar4.get() == 1: #GPU_mon
        if any(File.endswith(filetype_gmon) and File.startswith('GPU') for File in os.listdir(dir)):
            [_] = [fname for fname in dirdir if (fname.endswith(filetype_gmon) and fname.startswith('GPU'))]    
            gmonlog_path = os.path.join(dir, _)
            #try:
            n=0
            """ with open(gmonlog_path,'r') as f:
                ##################################
                #stripped = (line.strip(' ') for line in tmm_txt)
                for i in f.readlines():
                    #print(i)
                    if i.startswith(' ') or i.startswith('\t') or i == "\n" or i.startswith('GPU')or i.startswith('Tool'):#修改此處條件
                        n=n+1
                        #print(n)
                        
                    if i.startswith('date'):
                        break """
            with open(gmonlog_path,'r') as f:
                for i in f.readlines():
                    if i.startswith('date'):
                        break
                    else:
                        n=n+1

            
            data_gmon = pd.read_table(gmonlog_path, sep=",", skiprows=n ,encoding = "utf-8")#此行需要改
            #print(data_gmon)
            data_gmon.set_index('time', inplace=True)
            data_gmon1 = data_gmon.loc[:,data_gmon.columns.str.contains("1:t_gpu|1:mem_temp|1:gpc_clk|1:clk_mem|1:pwr_tgp|1:pwr_nvvdd|1:pwr_fbvdd|1:NVVDD|1:fan")]
            data_gmon1['1:pwr_tgp'] = data_gmon1['1:pwr_tgp']/1000
            data_gmon1['1:pwr_nvvdd'] = data_gmon1['1:pwr_nvvdd']/1000
            data_gmon1['1:pwr_fbvdd'] = data_gmon1['1:pwr_fbvdd']/1000
            gmon_last_avg = data_gmon1[-data_length_1:].mean().round(1)
            gmon_last_raw = data_gmon1
                
            #except:
            #    messagebox.showinfo("Notification","GPUmon data error.")
            #    return
            
            with pd.ExcelWriter(resultbook, mode="a", engine='openpyxl',if_sheet_exists="overlay") as writer:
                gmon_last_avg.to_excel(writer, sheet_name='avg', startcol = avg_col_num, header=False)
                gmon_last_raw.to_excel(writer, sheet_name="GPUMON_raw")
            avg_col_num = avg_col_num + 2
            """ book = load_workbook(resultbook)
            with pd.ExcelWriter(resultbook, engine='openpyxl') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                gmon_last_avg.to_excel(writer, sheet_name='avg',startrow=writer.sheets['avg'].max_row+1, header=False)
                gmon_last_raw.to_excel(writer, sheet_name='GPUmon_raw') """

        else:
            messagebox.showinfo("Notification","GPUmon data file does not exsist.")  
            return
    
    if CheckVar5.get() == 1: #HWinfo
        if any(File.endswith(filetype_hwi) for File in os.listdir(dir)):
            [_] = [fname for fname in dirdir if fname.endswith(filetype_hwi)]
            hwilog_path = os.path.join(dir, _)
            try:
                """ with open(hwilog_path) as infile:
                    reader = csv.reader(infile)
                    headers = next(reader)
                header_indices = [i for i, item in enumerate(headers) if item]
                data_hwi = pd.DataFrame(pd.read_csv(hwilog_path,encoding='ANSI',low_memory=False,usecols=header_indices))
                data_hwi.set_index('Time', inplace=True) """
                with open(hwilog_path) as infile:
                    reader = csv.reader(infile)
                    headers = next(reader)
                header_indices = [i for i, item in enumerate(headers) if item]
                data_hwi = pd.DataFrame(pd.read_csv(hwilog_path,encoding='ANSI',low_memory=False,usecols=header_indices))
                data_hwi.set_index('Time', inplace=True)
            
            except:
                messagebox.showinfo("Notification","HWInfo file log error, please check the raw data")
                return
            
            data_hwi1 = data_hwi.filter(regex="Drive Temperature|Write Rate|Read Rate",axis=1)
            #data_hwi2 = data_hwi.filter(regex="FAN",axis=1)
            data_hwinfo = data_hwi1
            data_hwinfo.columns = data_hwinfo.iloc[-1, 0:] + data_hwinfo.iloc[-2, 0:]  
            
            hwilog_path4 = data_hwinfo.iloc[0:-2,:].astype(float)
            hwi_last_avg = hwilog_path4.iloc[int(-data_length_1/3):].mean().round(1)
            hwi_last_raw = data_hwinfo.iloc[0:-2,:].astype(float)
            
            #hwi_last_raw = data_hwi.iloc[:-2,]#.astype(float)
            with pd.ExcelWriter(resultbook, mode="a", engine='openpyxl',if_sheet_exists="overlay") as writer:
                hwi_last_avg.to_excel(writer, sheet_name='avg', startcol = avg_col_num, header=False)
                hwi_last_raw.to_excel(writer, sheet_name="HWI_raw")
            avg_col_num = avg_col_num + 2

            """ book = load_workbook(resultbook)
            with pd.ExcelWriter(resultbook, engine='openpyxl') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                #hwi_last_avg.to_excel(writer, sheet_name='avg',startrow=writer.sheets['avg'].max_row, header=False)
                hwi_last_raw.to_excel(writer, sheet_name='HWI_raw') """

        else:
            messagebox.showinfo("Notification","HWinfo data file does not exsist.")  
            return


    """ book = load_workbook(resultbook)
    with pd.ExcelWriter(resultbook,engine='openpyxl')as writer:   
        writer.book = book
        writer.sheets.update = {ws.title: ws for ws in book.worksheets}
        worksheet = writer.sheets['avg']
        worksheet.column_dimensions['A'].width = 100
        for i in range(1,writer.sheets['avg'].max_row+1):
            worksheet.merge_cells('B%s:C%s' %(i,i)) """
    
    messagebox.showinfo('OK', "It's finish")



#button    
data_file = tk.StringVar()
data_file.set(dir)
widget = tk.Label(canvas, textvariable=data_file, font=('Microsoft New Tai Lue', 7),bg='Linen')
b = canvas.create_window(20, 20, anchor=NW, window=widget,width=420, height=40)
button1 = Button(canvas, text="Choose Folder", command=button_event1,font=('Microsoft New Tai Lue', 14))
button1.configure(width=12, activebackground="#33B5E5", relief=FLAT)
button1_window = canvas.create_window(450, 20, anchor=NW, window=button1, height=40)

button3 = Button(canvas, text="Run", font=(
                'Microsoft New Tai Lue', 14), command=run_all)#######################
button3.configure(width=12, activebackground="#33B5E5", relief=FLAT)
button3_window = canvas.create_window(40, 443, anchor=NW, window=button3)

#data length
def validate(P):
    if str.isdigit(P) or P == '':
        return True
    else:
        return False

mylabel = tk.Label(canvas, text='TAT acquiring data time (minutes): ')
mylabel_window = canvas.create_window(40, 80, anchor=NW, window=mylabel)
vcmd = (window.register(validate), '%P')
myentry = tk.Entry(canvas, validate='key', validatecommand=vcmd)
myentry.insert(0, "30")
canvas.create_window(250, 80, anchor=NW, window=myentry)

mylabel_2 = tk.Label(canvas, text='YOKO acquiring data time (minutes): ')
mylabel_window_2 = canvas.create_window(40, 110, anchor=NW, window=mylabel_2)
vcmd_2 = (window.register(validate), '%P')
myentry_2 = tk.Entry(canvas, validate='key', validatecommand=vcmd_2)
myentry_2.insert(0, "5")
canvas.create_window(250, 110, anchor=NW, window=myentry_2)

#data check bar
label1 = tk.Label(canvas, text='Intel TAT (PTAT) ', font=('Microsoft New Tai Lue', 14))
label1_window = canvas.create_window(40, 143, anchor=NW, window=label1)
CheckVar1 = IntVar(value=0)
tat_all = Checkbutton(text="TAT_all", variable=CheckVar1, font=('Microsoft New Tai Lue', 10),onvalue = 1)
tat_all.configure(width=10, activebackground="#33B5E5", relief=FLAT)
tat_all = canvas.create_window(40, 180, anchor=NW, window=tat_all)
CheckVar1_1 = IntVar(value=0)
dg_all = Checkbutton(text="DG", variable=CheckVar1_1, font=('Microsoft New Tai Lue', 10),onvalue = 1)
dg_all.configure(width=10, activebackground="#33B5E5", relief=FLAT)
dg_all = canvas.create_window(150, 180, anchor=NW, window=dg_all)

# tat_simple = Checkbutton(text="TAT_simple", variable=CheckVar1, font=('Microsoft New Tai Lue', 10),onvalue = 2)
# tat_simple.configure(width=10, activebackground="#33B5E5", relief=FLAT)
# tat_simple = canvas.create_window(150, 180, anchor=NW, window=tat_simple)

CheckVar2 = IntVar(value=0)
yoko_all = Checkbutton(text="Yokogawa", variable=CheckVar2, font=('Microsoft New Tai Lue', 14), anchor=W)
yoko_all.configure(width=10, activebackground="#33B5E5", relief=FLAT)
yoko_all = canvas.create_window(40, 223, anchor=NW, window=yoko_all)
# mylabel = tk.Label(canvas, text='How many channels? ')#以下為判斷多少CHANNEL
# mylabel_window = canvas.create_window(40, 263, anchor=NW, window=mylabel)
# entry_ch = tk.Entry(canvas,validate='key', validatecommand=vcmd)
# canvas.create_window(170, 265, anchor=NW, window=entry_ch)

CheckVar3 = IntVar(value=0)
tmm_all = Checkbutton(text="TMM", variable=CheckVar3, font=('Microsoft New Tai Lue', 14), anchor=W)
tmm_all.configure(width=10, activebackground="#33B5E5", relief=FLAT)
tmm_all = canvas.create_window(40, 309, anchor=NW, window=tmm_all)
# CheckVar3 = IntVar(value=1)
# hwi_all = Checkbutton(text="HWinfo", variable=CheckVar3, font=('Microsoft New Tai Lue', 14), anchor=W)
# hwi_all.configure(width=10, activebackground="#33B5E5", relief=FLAT)
# hwi_all = canvas.create_window(40, 309, anchor=NW, window=hwi_all)

CheckVar4 = IntVar(value=0)
gmon_all = Checkbutton(text="GPU_mon", variable=CheckVar4, font=('Microsoft New Tai Lue', 14), anchor=W)
gmon_all.configure(width=10, activebackground='#33B5E5', relief=FLAT)
canvas.create_window(40, 352, anchor=NW, window=gmon_all)

CheckVar5 = IntVar(value=0)
hwi_all = Checkbutton(text="HWinfo", variable=CheckVar5, font=('Microsoft New Tai Lue', 14), anchor=W)
hwi_all.configure(width=10, activebackground="#33B5E5", relief=FLAT)
hwi_all = canvas.create_window(40, 395, anchor=NW, window=hwi_all)

#window.iconbitmap('.\\pikachu.ico')

mainloop()